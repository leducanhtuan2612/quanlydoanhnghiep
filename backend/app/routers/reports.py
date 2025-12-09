from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from sqlalchemy import func, extract
from fastapi.responses import StreamingResponse

from app import models, database

from io import BytesIO
import openpyxl

# PDF
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4

router = APIRouter(prefix="/reports", tags=["Reports"])


# ============================================================
# 📊 SUMMARY REPORT (ĐÃ SỬA FULL, KHÔNG LỖI BIỂU ĐỒ)
# ============================================================
@router.get("/summary")
def get_summary(db: Session = Depends(database.get_db)):

    employees_count = db.query(models.Employee).count()
    customers_count = db.query(models.Customer).count()
    products_count = db.query(models.Product).count()

    inventory_items = db.query(models.Inventory).all()

    total_stock = sum(int(i.quantity or 0) for i in inventory_items)

    overview = {
        "employees_count": employees_count,
        "customers_count": customers_count,
        "products_count": products_count,
        "total_stock": total_stock,
    }

    # =====================================================
    #  🔧 FIX INVENTORY DATA — CHUẨN HÓA HOÀN TOÀN
    # =====================================================

    from collections import defaultdict
    inventory_map = defaultdict(int)

    for item in inventory_items:
        # Tên sản phẩm luôn là chuỗi an toàn
        name = item.product.name if item.product else "Unknown"

        # Chuẩn hóa quantity
        qty = item.quantity

        # convert to int safely
        try:
            qty = int(qty)
        except:
            qty = 0

        # Không cho âm
        if qty < 0:
            qty = 0

        # GỘP SẢN PHẨM TRÙNG
        inventory_map[name] += qty

    # Chuyển về mảng để Recharts sử dụng
    inventory_chart = [
        {"name": name, "stock": qty}
        for name, qty in inventory_map.items()
    ]

    # =====================================================
    #  ENTITY CHART
    # =====================================================
    entity_chart = [
        {"name": "Nhân viên", "value": employees_count},
        {"name": "Khách hàng", "value": customers_count},
        {"name": "Sản phẩm", "value": products_count},
    ]

    # TOP 5 tồn kho
    top_products = sorted(inventory_chart, key=lambda x: x["stock"], reverse=True)[:5]

    return {
        "overview": overview,
        "charts": {
            "inventory": inventory_chart,
            "entities": entity_chart,
        },
        "top_products": top_products,
    }

# ============================================================
# 💰 REVENUE REPORT
# ============================================================
@router.get("/revenue")
def get_revenue_report(db: Session = Depends(database.get_db)):

    completed_orders = db.query(models.Order).filter(models.Order.status == "Hoàn thành")

    by_month = (
        completed_orders.with_entities(
            extract("month", models.Order.date).label("month"),
            func.sum(models.Order.amount).label("total"),
        )
        .group_by(extract("month", models.Order.date))
        .order_by(extract("month", models.Order.date))
        .all()
    )

    by_month_data = [{"month": int(m[0]), "total": float(m[1] or 0)} for m in by_month]

# DOANH THU THEO DANH MỤC (ĐÃ SỬA)
    by_category = (
        completed_orders.join(models.Product)
        .with_entities(
            func.lower(models.Product.category).label("cat_norm"),
            func.sum(models.Order.amount).label("total"),
        )
        .group_by(func.lower(models.Product.category))
        .all()
    )

    by_category_data = [
        {
            "category": (cat or "khác").title(),
            "total": float(total or 0)
        }
        for cat, total in by_category
    ]


    by_region = (
        completed_orders.with_entities(
            models.Order.region, func.sum(models.Order.amount).label("total")
        )
        .group_by(models.Order.region)
        .all()
    )

    by_region_data = [
        {"region": (r[0] or "Không xác định"), "total": float(r[1] or 0)}
        for r in by_region
    ]

    total_revenue = sum(item["total"] for item in by_month_data)

    last_two_months = (
        completed_orders.with_entities(
            extract("month", models.Order.date).label("month"),
            func.sum(models.Order.amount).label("total"),
        )
        .group_by(extract("month", models.Order.date))
        .order_by(extract("month", models.Order.date).desc())
        .limit(2)
        .all()
    )

    growth = 0
    if len(last_two_months) == 2:
        cur = float(last_two_months[0][1] or 0)
        prev = float(last_two_months[1][1] or 0)
        growth = (cur - prev) / prev * 100 if prev > 0 else 0

    return {
        "total_revenue": total_revenue,
        "growth": growth,
        "by_month": by_month_data,
        "by_category": by_category_data,
        "by_region": by_region_data,
    }


# ============================================================
# 🏆 TOP PRODUCTS
# ============================================================
@router.get("/top-products")
def get_top_products(db: Session = Depends(database.get_db)):

    result = (
        db.query(
            models.Product.name.label("product"),
            func.sum(models.Order.quantity).label("total_sold"),
            func.sum(models.Order.amount).label("revenue"),
        )
        .join(models.Product, models.Product.id == models.Order.product_id)
        .filter(models.Order.status == "Hoàn thành")
        .group_by(models.Product.id)
        .order_by(func.sum(models.Order.quantity).desc())
        .limit(10)
        .all()
    )

    return [
        {
            "product": r.product,
            "total_sold": int(r.total_sold or 0),
            "revenue": float(r.revenue or 0),
        }
        for r in result
    ]


# ============================================================
# 👤 TOP CUSTOMERS
# ============================================================
@router.get("/top-customers")
def get_top_customers(db: Session = Depends(database.get_db)):

    result = (
        db.query(
            models.Customer.name.label("customer"),
            func.count(models.Order.id).label("order_count"),
            func.sum(models.Order.amount).label("total_spent"),
        )
        .join(models.Customer, models.Customer.id == models.Order.customer_id)
        .filter(models.Order.status == "Hoàn thành")
        .group_by(models.Customer.id)
        .order_by(func.sum(models.Order.amount).desc())
        .limit(10)
        .all()
    )

    return [
        {
            "customer": r.customer,
            "order_count": int(r.order_count or 0),
            "total_spent": float(r.total_spent or 0),
        }
        for r in result
    ]


# ============================================================
# 📤 EXPORT EXCEL – FULL DATA
# ============================================================
@router.get("/export/excel")
def export_excel(db: Session = Depends(database.get_db)):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BAO CAO DOANH THU"

    # ===== TITLE =====
    ws.merge_cells("A1:E1")
    ws["A1"] = "BÁO CÁO DOANH THU TỔNG HỢP"
    ws["A1"].font = openpyxl.styles.Font(size=18, bold=True)
    ws.append([])

    # =======================================
    # 1️⃣ LẤY DỮ LIỆU DOANH THU THEO THÁNG
    # =======================================
    completed_orders = db.query(models.Order).filter(
        models.Order.status == "Hoàn thành"
    )

    by_month = (
        completed_orders.with_entities(
            extract("month", models.Order.date).label("month"),
            func.sum(models.Order.amount).label("total"),
        )
        .group_by(extract("month", models.Order.date))
        .order_by(extract("month", models.Order.date))
        .all()
    )

    # Tổng doanh thu
    total_revenue = sum(float(r[1] or 0) for r in by_month)

    # Tính tăng trưởng
    last_two = list(by_month)[-2:]
    growth = 0
    if len(last_two) == 2:
        prev = float(last_two[0][1] or 0)
        cur = float(last_two[1][1] or 0)
        if prev > 0:
            growth = (cur - prev) / prev * 100

    # ===== WRITE SUMMARY =====
    ws.append(["Tổng doanh thu", total_revenue])
    ws.append(["Tăng trưởng so với tháng trước (%)", round(growth, 2)])
    ws.append([])
    ws.append(["Doanh thu theo tháng"])
    ws.append(["Tháng", "Doanh thu (VND)"])

    for m, total in by_month:
        ws.append([int(m), float(total or 0)])

    ws.append([])
    ws.append([])

    # =======================================
    # 2️⃣ DOANH THU THEO DANH MỤC
    # =======================================
    ws.append(["Doanh thu theo danh mục"])
    ws.append(["Danh mục", "Doanh thu"])

    by_category = (
        completed_orders.join(models.Product)
        .with_entities(
            models.Product.category,
            func.sum(models.Order.amount),
        )
        .group_by(models.Product.category)
        .all()
    )

    for category, total in by_category:
        ws.append([category or "Khác", float(total or 0)])

    ws.append([])
    ws.append([])

    # =======================================
    # 3️⃣ DOANH THU THEO KHU VỰC
    # =======================================
    ws.append(["Doanh thu theo khu vực"])
    ws.append(["Khu vực", "Doanh thu"])

    by_region = (
        completed_orders.with_entities(
            models.Order.region,
            func.sum(models.Order.amount)
        )
        .group_by(models.Order.region)
        .all()
    )

    for region, total in by_region:
        ws.append([region or "Không xác định", float(total or 0)])

    ws.append([])
    ws.append([])

    # =======================================
    # 4️⃣ TOP 10 SẢN PHẨM BÁN CHẠY
    # =======================================
    ws.append(["Top 10 sản phẩm bán chạy"])
    ws.append(["Sản phẩm", "Số lượng bán", "Doanh thu"])

    top_products = (
        db.query(
            models.Product.name.label("product"),
            func.sum(models.Order.quantity).label("sold"),
            func.sum(models.Order.amount).label("revenue"),
        )
        .join(models.Product)
        .filter(models.Order.status == "Hoàn thành")
        .group_by(models.Product.id)
        .order_by(func.sum(models.Order.quantity).desc())
        .limit(10)
        .all()
    )

    for p in top_products:
        ws.append([p.product, int(p.sold or 0), float(p.revenue or 0)])

    ws.append([])
    ws.append([])

    # =======================================
    # 5️⃣ TOP 10 KHÁCH HÀNG MUA NHIỀU NHẤT
    # =======================================
    ws.append(["Top 10 khách hàng mua nhiều nhất"])
    ws.append(["Khách hàng", "Số đơn", "Tổng chi tiêu"])

    top_customers = (
        db.query(
            models.Customer.name,
            func.count(models.Order.id),
            func.sum(models.Order.amount),
        )
        .join(models.Customer)
        .filter(models.Order.status == "Hoàn thành")
        .group_by(models.Customer.id)
        .order_by(func.sum(models.Order.amount).desc())
        .limit(10)
        .all()
    )

    for name, count_order, spending in top_customers:
        ws.append([name, int(count_order or 0), float(spending or 0)])

    # ===== Resize columns =====
    for col in ["A", "B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 25

    # ===== SAVE =====
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": "attachment; filename=bao_cao_doanh_thu.xlsx"},
    )

# ============================================================
# 📄 EXPORT PDF FULL – KHÔNG LỖI FONT
# ============================================================
@router.get("/export/pdf")
def export_pdf(db: Session = Depends(database.get_db)):

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)

    # ========== 1️⃣ LẤY DỮ LIỆU CHUNG ==========
    completed_orders = db.query(models.Order).filter(models.Order.status == "Hoàn thành")

    total_revenue = (
        completed_orders.with_entities(func.sum(models.Order.amount)).scalar() or 0
    )

    # Tăng trưởng
    last_two = (
        completed_orders.with_entities(
            extract("month", models.Order.date),
            func.sum(models.Order.amount)
        )
        .group_by(extract("month", models.Order.date))
        .order_by(extract("month", models.Order.date).desc())
        .limit(2)
        .all()
    )

    growth = 0
    if len(last_two) == 2:
        cur, prev = float(last_two[0][1] or 0), float(last_two[1][1] or 0)
        if prev > 0:
            growth = ((cur - prev) / prev) * 100

    # ========== HEADER ==========
    p.setFont("Helvetica-Bold", 18)
    p.drawString(50, 800, "BAO CAO DOANH THU")

    p.setFont("Helvetica", 12)
    p.drawString(50, 770, f"Tong doanh thu: {total_revenue:,.0f} VND")
    p.drawString(50, 750, f"Tang truong thang truoc: {growth:.2f}%")

    y = 720

    # ========== 2️⃣ BẢNG DOANH THU THEO THÁNG ==========
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, "Doanh thu theo thang:")
    y -= 25

    by_month = (
        completed_orders.with_entities(
            extract("month", models.Order.date),
            func.sum(models.Order.amount),
        )
        .group_by(extract("month", models.Order.date))
        .order_by(extract("month", models.Order.date))
        .all()
    )

    p.setFont("Helvetica", 12)
    for m, total in by_month:
        p.drawString(60, y, f"- Thang {int(m)}: {float(total):,.0f} VND")
        y -= 20

    y -= 10

    # ========== 3️⃣ DOANH THU THEO DANH MUC ==========
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, "Doanh thu theo danh muc:")
    y -= 25

    by_category = (
        completed_orders.join(models.Product)
        .with_entities(models.Product.category, func.sum(models.Order.amount))
        .group_by(models.Product.category)
        .all()
    )

    p.setFont("Helvetica", 12)
    for cat, total in by_category:
        cat = cat or "Khac"
        p.drawString(60, y, f"- {cat}: {float(total):,.0f} VND")
        y -= 20

    y -= 10

    # ========== 4️⃣ DOANH THU THEO KHU VUC ==========
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, "Doanh thu theo khu vuc:")
    y -= 25

    by_region = (
        completed_orders.with_entities(models.Order.region, func.sum(models.Order.amount))
        .group_by(models.Order.region)
        .all()
    )

    p.setFont("Helvetica", 12)
    for region, total in by_region:
        region = region or "Khong xac dinh"
        p.drawString(60, y, f"- {region}: {float(total):,.0f} VND")
        y -= 20

    y -= 10

    # ========== 5️⃣ TOP SAN PHAM ==========
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, "Top 10 san pham ban chay:")
    y -= 25

    top_products = (
        db.query(
            models.Product.name,
            func.sum(models.Order.quantity),
            func.sum(models.Order.amount)
        )
        .join(models.Product)
        .filter(models.Order.status == "Hoàn thành")
        .group_by(models.Product.id)
        .order_by(func.sum(models.Order.quantity).desc())
        .limit(10)
        .all()
    )

    p.setFont("Helvetica", 12)
    for name, qty, revenue in top_products:
        p.drawString(60, y, f"- {name}: {int(qty or 0)} SP, {float(revenue or 0):,.0f} VND")
        y -= 20

    y -= 10

    # ========== 6️⃣ TOP KHACH HANG ==========
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, "Top 10 khach hang chi tieu nhieu nhat:")
    y -= 25

    top_customers = (
        db.query(
            models.Customer.name,
            func.count(models.Order.id),
            func.sum(models.Order.amount)
        )
        .join(models.Customer)
        .filter(models.Order.status == "Hoàn thành")
        .group_by(models.Customer.id)
        .order_by(func.sum(models.Order.amount).desc())
        .limit(10)
        .all()
    )

    p.setFont("Helvetica", 12)
    for name, count_order, spend in top_customers:
        p.drawString(60, y, f"- {name}: {int(count_order)} don, {float(spend):,.0f} VND")
        y -= 20

    # ===== FINISH PAGE =====
    p.showPage()
    p.save()

    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=bao_cao_doanh_thu.pdf"},
    )
# ==========================================================
# 📊 REPORT: SỐ ĐƠN THEO TRẠNG THÁI
# ==========================================================
@router.get("/report/status")
def order_report_status(db: Session = Depends(database.get_db)):

    data = (
        db.query(
            models.Order.status,
            func.count(models.Order.id).label("count")
        )
        .group_by(models.Order.status)
        .all()
    )

    return [
        {
            "status": status or "Không xác định",
            "count": int(count or 0)       # ⭐ MUST FIX – ép int
        }
        for status, count in data
    ]

# ==========================================================
# 📅 REPORT: SỐ ĐƠN THEO THÁNG
# ==========================================================
@router.get("/report/month")
def order_report_month(db: Session = Depends(database.get_db)):

    data = (
        db.query(
            extract("month", models.Order.date).label("month"),
            func.count(models.Order.id).label("count")
        )
        .group_by("month")
        .order_by("month")
        .all()
    )

    return [
        {
            "month": int(month),
            "count": int(count or 0)        # ⭐ MUST FIX – ép int
        }
        for month, count in data
    ]
# ============================================================
# 📤 EXPORT EXCEL – SUMMARY REPORT
# ============================================================
@router.get("/export/summary-excel")
def export_summary_excel(db: Session = Depends(database.get_db)):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BAO CAO TONG HOP"

    ws.merge_cells("A1:D1")
    ws["A1"] = "BÁO CÁO TỔNG HỢP HỆ THỐNG"
    ws["A1"].font = openpyxl.styles.Font(size=18, bold=True)
    ws.append([])

    # ========== LẤY DỮ LIỆU ==========

    employees_count = db.query(models.Employee).count()
    customers_count = db.query(models.Customer).count()
    products_count = db.query(models.Product).count()
    inventory_items = db.query(models.Inventory).all()

    total_stock = sum((i.quantity or 0) for i in inventory_items)

    ws.append(["Thông tin", "Giá trị"])
    ws.append(["Tổng nhân viên", employees_count])
    ws.append(["Tổng khách hàng", customers_count])
    ws.append(["Tổng sản phẩm", products_count])
    ws.append(["Tổng tồn kho", total_stock])

    ws.append([])
    ws.append(["TỒN KHO THEO SẢN PHẨM"])
    ws.append(["Sản phẩm", "Tồn kho"])

    for i in inventory_items:
        ws.append([
            i.product.name if i.product else "Unknown",
            int(i.quantity or 0)
        ])

    ws.append([])
    ws.append(["TOP 5 SẢN PHẨM TỒN NHIỀU"])
    ws.append(["Sản phẩm", "Tồn kho"])

    inventory_sorted = sorted(
        inventory_items, key=lambda x: x.quantity or 0, reverse=True
    )[:5]

    for i in inventory_sorted:
        ws.append([i.product.name, int(i.quantity or 0)])

    # Resize columns
    for col in ["A", "B", "C", "D"]:
        ws.column_dimensions[col].width = 25

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": "attachment; filename=bao_cao_tong_hop.xlsx"
        },
    )
# ============================================================
# 📄 EXPORT PDF – SUMMARY REPORT (FULL, KHONG DAU)
# ============================================================
@router.get("/export/summary-pdf")
def export_summary_pdf(db: Session = Depends(database.get_db)):

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)

    # FONT
    p.setFont("Helvetica-Bold", 18)
    p.drawString(50, 800, "BAO CAO TONG HOP HE THONG")

    p.setFont("Helvetica", 12)

    # ===============================
    # 1️⃣ LAY DU LIEU
    # ===============================
    employees = db.query(models.Employee).count()
    customers = db.query(models.Customer).count()
    products = db.query(models.Product).count()
    inventory_items = db.query(models.Inventory).all()
    total_stock = sum((i.quantity or 0) for i in inventory_items)

    # Thong ke don hang
    orders_by_status = (
        db.query(models.Order.status, func.count(models.Order.id))
        .group_by(models.Order.status)
        .all()
    )

    orders_by_month = (
        db.query(
            extract("month", models.Order.date).label("month"),
            func.count(models.Order.id)
        )
        .group_by("month")
        .order_by("month")
        .all()
    )

    # ===============================
    # 2️⃣ TONG QUAN
    # ===============================
    y = 760
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, "1. Tong quan he thong")
    y -= 25

    p.setFont("Helvetica", 12)
    p.drawString(60, y, f"- Tong nhan vien: {employees}")
    y -= 20
    p.drawString(60, y, f"- Tong khach hang: {customers}")
    y -= 20
    p.drawString(60, y, f"- Tong san pham: {products}")
    y -= 20
    p.drawString(60, y, f"- Tong so luong ton kho: {total_stock}")
    y -= 30

    # ===============================
    # 3️⃣ TOP SAN PHAM TON
    # ===============================
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, "2. Top 5 san pham ton kho nhieu nhat")
    y -= 25

    top5 = sorted(inventory_items, key=lambda x: x.quantity or 0, reverse=True)[:5]

    p.setFont("Helvetica", 12)
    for item in top5:
        p.drawString(
            60, y,
            f"- {item.product.name if item.product else 'Unknown'}: {int(item.quantity or 0)}"
        )
        y -= 20

    y -= 20

    # ===============================
    # 4️⃣ DANH SACH TON KHO DAY DU
    # ===============================
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, "3. Danh sach ton kho tat ca san pham")
    y -= 25

    p.setFont("Helvetica", 11)

    for item in inventory_items:
        text = f"- {item.product.name if item.product else 'Unknown'}: {int(item.quantity or 0)}"
        p.drawString(60, y, text)
        y -= 15
        if y < 40:       # auto xuống trang
            p.showPage()
            y = 800
            p.setFont("Helvetica", 11)

    # ===============================
    # 5️⃣ SO DON THEO TRANG THAI
    # ===============================
    if y < 120:
        p.showPage()
        y = 800

    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, "4. So don theo trang thai")
    y -= 25

    p.setFont("Helvetica", 12)
    for status, count in orders_by_status:
        p.drawString(60, y, f"- {status}: {int(count)} don")
        y -= 20

    y -= 20

    # ===============================
    # 6️⃣ SO DON THEO THANG
    # ===============================
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, "5. So don theo thang")
    y -= 25

    for month, count in orders_by_month:
        p.drawString(60, y, f"- Thang {int(month)}: {int(count)} don")
        y -= 20

    y -= 30

    # ===============================
    # 7️⃣ GHI CHU HE THONG
    # ===============================
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, "6. Ghi chu he thong")
    y -= 25

    p.setFont("Helvetica", 12)
    p.drawString(60, y, "- Du lieu duoc tong hop tu he thong quan ly doanh nghiep Tuấn ERP.")
    y -= 20
    p.drawString(60, y, "- Bao cao duoc xuat tu module Reports.")
    y -= 20
    p.drawString(60, y, "- Muc dich su dung: quan tri, lam bao cao, phan tich hoat dong.")
    y -= 20

    p.showPage()
    p.save()

    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=bao_cao_tong_hop.pdf"},
    )
# app/routers/reports.py
@router.get("/forecast")
def forecast_revenue():
    # Demo: dự đoán dựa trên tăng trưởng 3 tháng gần nhất
    import random
    
    months = [9, 10, 11]
    real = [400_000, 620_000, 900_000]

    # Dự đoán tháng 12
    next_value = int(real[-1] * random.uniform(1.05, 1.25))

    return {
        "real": [
            {"month": m, "value": real[i]}
            for i, m in enumerate(months)
        ],
        "forecast": [
            {"month": 12, "value": next_value}
        ],
        "summary": {
            "predicted_revenue": next_value,
            "growth_rate": round((next_value - real[-1]) / real[-1] * 100, 1),
            "suggestion": "Nên nhập thêm nhóm hàng bán chạy để tăng trưởng tốt hơn."
        }
    }
